# /// script
# requires-python = ">=3.9"
# dependencies = ["openpyxl", "pyyaml", "rdflib"]
# ///
"""Transform a FutuRaM composition dataset (CSV/Excel) into composition RDF. Level
from `parameterCode` (c-p/m-c/e-m); ABSOLUTE kg/unit -> per-layer kg/kg. Each
`productionYear` is a product INSTANCE; the m->e layer is YEAR-INVARIANT chemistry.
"""
from __future__ import annotations

import argparse
import csv
import pathlib
import sys
from collections import defaultdict

import yaml

from etl import limit_corrections as _lc

# --- the one universal lookup: chemical symbol -> element name (periodic table).
# NOT dataset-specific. Every symbol seen in `element` must resolve here (or be a
# '_'-joined compound of these, or the otherOrUndefined literal) else we fail.
SYMBOL_TO_ELEMENT = {
    "H": "Hydrogen", "He": "Helium", "Li": "Lithium", "Be": "Beryllium",
    "B": "Boron", "C": "Carbon", "N": "Nitrogen", "O": "Oxygen",
    "F": "Fluorine", "Ne": "Neon", "Na": "Sodium", "Mg": "Magnesium",
    "Al": "Aluminium", "Si": "Silicon", "P": "Phosphorus", "S": "Sulfur",
    "Cl": "Chlorine", "Ar": "Argon", "K": "Potassium", "Ca": "Calcium",
    "Sc": "Scandium", "Ti": "Titanium", "V": "Vanadium", "Cr": "Chromium",
    "Mn": "Manganese", "Fe": "Iron", "Co": "Cobalt", "Ni": "Nickel",
    "Cu": "Copper", "Zn": "Zinc", "Ga": "Gallium", "Ge": "Germanium",
    "As": "Arsenic", "Se": "Selenium", "Br": "Bromine", "Kr": "Krypton",
    "Rb": "Rubidium", "Sr": "Strontium", "Y": "Yttrium", "Zr": "Zirconium",
    "Nb": "Niobium", "Mo": "Molybdenum", "Tc": "Technetium", "Ru": "Ruthenium",
    "Rh": "Rhodium", "Pd": "Palladium", "Ag": "Silver", "Cd": "Cadmium",
    "In": "Indium", "Sn": "Tin", "Sb": "Antimony", "Te": "Tellurium",
    "I": "Iodine", "Xe": "Xenon", "Cs": "Caesium", "Ba": "Barium",
    "La": "Lanthanum", "Ce": "Cerium", "Pr": "Praseodymium", "Nd": "Neodymium",
    "Pm": "Promethium", "Sm": "Samarium", "Eu": "Europium", "Gd": "Gadolinium",
    "Tb": "Terbium", "Dy": "Dysprosium", "Ho": "Holmium", "Er": "Erbium",
    "Tm": "Thulium", "Yb": "Ytterbium", "Lu": "Lutetium", "Hf": "Hafnium",
    "Ta": "Tantalum", "W": "Tungsten", "Re": "Rhenium", "Os": "Osmium",
    "Ir": "Iridium", "Pt": "Platinum", "Au": "Gold", "Hg": "Mercury",
    "Tl": "Thallium", "Pb": "Lead", "Bi": "Bismuth", "Po": "Polonium",
    "At": "Astatine", "Rn": "Radon", "Fr": "Francium", "Ra": "Radium",
    "Ac": "Actinium", "Th": "Thorium", "Pa": "Protactinium", "U": "Uranium",
}

# The element-cell remainder literal (mass not assigned to a named element): the
# ELEMENT-level unknown filler, mapped to futuram:unknownElement (uniform with
# unknownMaterial/Component).
OTHER_ELEMENT = "otherOrUndefinedElements"   # the source literal we detect
UNKNOWN_ELEMENT = "unknownElement"           # … and the class we map it to

# Every measured value is a RECTANGULAR distribution whose half-width is derived
# from the DQ scores by futuram:FuturamDQS. The ETL stores only the DQ scores + a
# bare RectangularDistribution, ignoring the CSV's calculated uncertainty% column.

# parameterCode -> (whole-level, part-level). The ONLY structural knowledge,
# and it is the dataset's own documented coding, not a name guess.
LEVEL_OF_CODE = {
    "c-p": ("Product", "Component"),
    "m-c": ("Component", "Material"),
    "e-m": ("Material", "Element"),
}

# column families that spell out an identity path (leaf = the node's class).
COMPONENT_KEY_COLS = ["componentKeyLevel0", "componentKeyLevel1",
                      "componentKeyLevel2"]
MATERIAL_KEY_COLS = ["materialKeyLevel0", "materialKeyLevel1",
                     "materialKeyLevel2", "materialKeyLevel3",
                     "materialKeyLevel4"]
PRODUCT_KEY_COL = "productKeyLevel2"

# Human-readable drivetrain names (from productKeyLevel1, the drivetrain axis).
_DRIVETRAIN_NAME = {
    "elvBEV": "battery-electric vehicle (BEV)",
    "elvPHEV": "plug-in hybrid vehicle (PHEV)",
    "elvHEV": "hybrid vehicle (HEV)",
    "elvPetrol": "petrol vehicle",
    "elvDiesel": "diesel vehicle",
    "elvOther": "other-drivetrain vehicle",
}
# Trailing 4 digits of a V03..-style productKeyLevel2 encode the body segment:
# 0000=standard, 01xx=passenger segments A-F, 02xx=the JA-JF family (EU Euro Car
# Segment size/role descriptions).
_SEGMENT_LETTER = {  # last 4 digits -> (segment code, family description)
    "0000": ("standard", "standard / segment-unspecified (the segment-averaged car)"),
    "0101": ("A", "segment A (mini / city car — the smallest passenger cars)"),
    "0102": ("B", "segment B (small car / supermini — small hatchbacks)"),
    "0103": ("C", "segment C (medium / lower-medium — small-family / compact)"),
    "0104": ("D", "segment D (large / upper-medium — large-family cars)"),
    "0105": ("E", "segment E (executive — large saloons)"),
    "0106": ("F", "segment F (luxury — full-size luxury cars)"),
    "0201": ("JA", "segment JA (small SUV / crossover — J-family A)"),
    "0202": ("JB", "segment JB (compact SUV — J-family B)"),
    "0203": ("JC", "segment JC (mid-size SUV — J-family C)"),
    "0204": ("JD", "segment JD (large SUV — J-family D)"),
    "0205": ("JE", "segment JE (executive SUV — J-family E)"),
    "0206": ("JF", "segment JF (large / luxury SUV — J-family F)"),
}


def _product_label(code: str, drivetrain: str | None) -> str | None:
    """A rich rdfs:label for a vehicle product CLASS, decoding the V-code: drivetrain
    (productKeyLevel1) + body segment (trailing 4 digits). The raw code is NOT
    appended (it is the IRI's local name). Returns None for an undecodable code."""
    if not code or not code.startswith("V") or len(code) < 4:
        return None
    seg = _SEGMENT_LETTER.get(code[-4:])
    dt_name = _DRIVETRAIN_NAME.get(str(drivetrain or ""), str(drivetrain or "").removeprefix("elv") or None)
    if seg is None and dt_name is None:
        return None
    parts = []
    if dt_name:
        parts.append(dt_name)
    if seg:
        parts.append(seg[1])
    return " — ".join(parts)

# data-quality columns -> DQV dimension localnames (the scenario `quality` map).
DQ_COLS = {
    "dqAccuracy": "Accuracy", "dqValidity": "Validity",
    "dqConsistency": "Consistency", "dqIntegrity": "Integrity",
    "dqTimeliness": "Timeliness", "dqCompleteness": "Completeness",
}

# Every composition statement MUST carry all six DQ dimensions (the uncertainty rule
# needs the complete vector); common.vocab.fill_quality fills any missing dimension
# from the shared DEFAULT_QUALITY, so the invariant holds everywhere.
from common.vocab import fill_quality as _fill_quality


# ---------------------------------------------------------------------------
# Reading (Excel or CSV) into a list of dict rows keyed by header name
# ---------------------------------------------------------------------------

def read_rows(path: pathlib.Path, sheet=None):
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        return [dict(zip(header, r)) for r in it]
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def num(v):
    """Parse a numeric cell (Excel gives floats; CSV gives strings)."""
    if v is None or v == "":
        return None
    return float(v)


# cells that mean "no value" (empty / placeholders); excluded from key paths.
_ABSENT = {"", "n/a", "na", "none", "nan", "null", "-"}


def _present(v):
    return v is not None and str(v).strip().lower() not in _ABSENT


def key_path(row, cols):
    """The non-empty identity path for a node (e.g. component or material),
    skipping empty/placeholder cells (N/A, nan, ...)."""
    return tuple(str(row[c]).strip() for c in cols if c in row and _present(row[c]))


# catch-all / unattributed source categories map to the generic unknown filler
# classes (the resolver reconciles them). Detected by a data-driven token pattern
# ('other'/'undefined'/'unknown'/'mixed'/'rest'), not a hardcoded name list.
_CATCHALL_TOKENS = ("other", "undefined", "unknown", "unspecified", "mixed",
                    "misc", "rest", "n/a", "na")


def is_catchall(key):
    k = str(key).strip().lower()
    return any(tok in k for tok in _CATCHALL_TOKENS)


# A resolved CLASS localname the ETL must NOT author as a node (otherOrUndefined*,
# elvRest, unknown<level>): OMITTED so the resolver re-infers the remainder. Precise
# prefix/exact match so real classes (elvOther, otherCuAlloys, Vanadium) survive.
def _is_omit_class(cls_localname):
    s = str(cls_localname)
    return (s.startswith("otherOrUndefined") or s.startswith("unknown")
            or s == "elvRest")


# A few source leaf KEYS spell a concept the ontology declares under a different
# localname (plural/short form); alias to the EXISTING class, checked before
# catch-all/fail-loud so a known concept is never demoted to the unknown filler.
_KEY_ALIASES = {
    "catalysts": "catalystsMaterial",
    # The source CSVs spell this with one 'l' ("elvDepolutionComponents"); the
    # ontology declares "elvDepollutionComponents" (two 'l's). Alias the source
    # typo so the row resolves (CSVs are upstream-owned, must be tolerated).
    "elvDepolutionComponents": "elvDepollutionComponents",
}


# Some source keys are DECLARED ontology classes yet are semantically their level's
# remainder catch-all (e.g. elvRest); they map to the unknown<Level> filler (via
# class_for's `unknown_cls`), OVERRIDING "declared class wins" below.
_CATCHALL_CLASSES = {"elvRest"}


# ---------------------------------------------------------------------------
# Element symbol -> futuram class
# ---------------------------------------------------------------------------

def element_class(symbol, ontology_classes):
    """Map an `element` cell to a futuram Element class localname. A compound tag
    ('Cr_Mo') is kept as one node classed by its first constituent. Every plain
    symbol must be in the periodic table AND an ontology class, else we raise."""
    # KNOWN-SYMBOL-WINS: resolve a real chemical symbol BEFORE any catch-all
    # check, so a symbol whose name happens to contain a catch-all token (e.g.
    # 'Na' contains "na") is never demoted to the unknown filler.
    parts = symbol.split("_")          # compound co-measurement, e.g. Cr_Mo
    head = parts[0]
    if head in SYMBOL_TO_ELEMENT:
        cls = SYMBOL_TO_ELEMENT[head]
        if cls not in ontology_classes:
            raise ValueError(
                f"element {symbol!r} -> futuram:{cls} is not a class in the ontology")
        return cls, symbol
    # not a known element: the remainder literal or any catch-all placeholder, both
    # mapped to futuram:unknownElement. The node name keeps the source symbol so
    # distinct remainders stay distinct nodes; only the CLASS is the generic filler.
    if symbol == OTHER_ELEMENT or is_catchall(symbol):
        cls = UNKNOWN_ELEMENT if UNKNOWN_ELEMENT in ontology_classes else None
        return cls, symbol
    raise ValueError(
        f"element symbol {symbol!r}: '{head}' is not a known chemical symbol "
        f"(add it to SYMBOL_TO_ELEMENT) and is not a recognised catch-all")
    return cls, symbol


# ---------------------------------------------------------------------------
# Build the per-instance composition tree and normalise to kg/kg per layer
# ---------------------------------------------------------------------------


def ontology_class_set():
    """Localnames of every owl:Class in the TBox + hierarchy (to validate that a
    key really is a futuram class)."""
    from rdflib import Graph, RDF, OWL
    here = pathlib.Path(__file__).resolve()
    for cand in (here, *here.parents):       # repo root = nearest pyproject.toml
        if (cand / "pyproject.toml").exists():
            here = cand
            break
    g = Graph()
    g.parse(here / "ontology" / "tbox" / "composition-statement.ttl", format="turtle")
    g.parse(here / "ontology" / "tbox" / "futuram-hierarchy.ttl", format="turtle")
    return {str(s).split("#")[-1] for s in g.subjects(RDF.type, OWL.Class)}


def node_id(prefix, path, year):
    """A stable, readable node name unique within an instance (year-suffixed)."""
    leaf = path[-1] if path else prefix
    return f"{leaf}_{year}"


# Every measured value is a RECTANGULAR distribution whose half-width is DERIVED
# from the DQ scores (uncertaintyLimitStrategy). The ETL reads ONLY the DQ scores,
# never the CSV's calculated uncertainty% column.
def aggregate_rows(rs):
    """mean/quality over duplicate rows of one (whole,part) key. Returns the mean
    value and the FULL six-dimension DQV quality dict (spread is derived from quality
    downstream); any dimension the source omits is filled from DEFAULT_QUALITY."""
    vals = [num(r["value"]) for r in rs if num(r["value"]) is not None]
    mean = sum(vals) / len(vals)
    q = {dim: float(rs[0][col]) for col, dim in DQ_COLS.items()
         if rs[0].get(col) not in (None, "")}
    return mean, _fill_quality(q)        # always the complete six-dimension vector


def _collect_facts(rows):
    """Group one instance's rows by structural key: m-c by (comp, mat), e-m by
    (comp, mat, element), c-p totals per comp (cross-check only)."""
    mc_by = defaultdict(list)   # (comp_path, mat_path) -> rows
    em_by = defaultdict(list)   # (comp_path, mat_path, element) -> rows
    cp_total = defaultdict(float)   # comp_path -> sum of c-p values (cross-check)
    for r in rows:
        code = r.get("parameterCode")
        if code not in LEVEL_OF_CODE:
            continue
        comp = key_path(r, COMPONENT_KEY_COLS)
        mat = key_path(r, MATERIAL_KEY_COLS)
        if code == "m-c":
            mc_by[(comp, mat)].append(r)
        elif code == "e-m":
            em_by[(comp, mat, r["element"])].append(r)
        elif code == "c-p" and num(r.get("value")) is not None:
            cp_total[comp] += num(r["value"])
    return mc_by, em_by, cp_total


# Max relative drift an element fraction may show across years before the transform
# REFUSES the dataset (observed real drift is rounding noise); a genuine composition
# change is a different material, not smuggled in via the year.
EM_YEAR_TOL = 0.01


def _em_canonical(by_inst):
    """The year-invariant material->element table, ENFORCED: per (product, component,
    material) the e-m kg/kg fractions must match across years (canonical = per-element
    mean, raises beyond EM_YEAR_TOL). Returns {(product,comp,mat): {sym: (best, q)}}."""
    per_year = defaultdict(dict)   # (prod, comp, mat) -> {year: {sym: (best, q)}}
    for (prod, year), rows in by_inst.items():
        mc_by, em_by, _ = _collect_facts(rows)
        mat_mass = {k: aggregate_rows(rs)[0] for k, rs in mc_by.items()}
        for (comp, mat, sym), rs in em_by.items():
            mm = mat_mass.get((comp, mat))
            if not mm:
                continue
            val, q = aggregate_rows(rs)
            per_year[(prod, comp, mat)].setdefault(year, {})[sym] = (val / mm, q)

    canon, offenders = {}, []
    for (prod, comp, mat), years in per_year.items():
        table = {}
        for sym in sorted({s for d in years.values() for s in d}):
            # a year where the material exists but the element is absent reads
            # as 0.0 — present-vs-absent across years IS a composition change.
            bests = [d.get(sym, (0.0,))[0] for d in years.values()]
            top = max(bests)
            spread = (top - min(bests)) / top if top > 0 else 0.0
            if spread > EM_YEAR_TOL:
                offenders.append(
                    f"{prod} / {comp[-1] if comp else '?'} / "
                    f"{mat[-1] if mat else '?'} / {sym}: "
                    f"{min(bests):.6g}..{top:.6g} kg/kg across years "
                    f"{min(years)}..{max(years)} (rel spread {spread:.3g})")
                continue
            entries = [d[sym] for d in years.values() if sym in d]
            n = len(entries)
            best = sum(e[0] for e in entries) / n
            if best <= 0.0:
                continue        # zero-content element: omitted, like emit()
            table[sym] = (best, entries[0][1])             # (best, quality)
        canon[(prod, comp, mat)] = table
    if offenders:
        raise ValueError(
            "material->element composition is NOT year-invariant (a material's "
            "element make-up is intrinsic chemistry; fix the source data, or "
            "model a real composition change as a distinct material):\n  "
            + "\n  ".join(offenders))
    return canon


def build_instance(rows, year, ontology_classes, em_canon=None,
                   em_emitted=None, scope=None, node_time=None, corrections=None):
    """Build (nodes, statements, info) for one production-year (or -period) instance.
    Product/Component instances are typed by their TIMELESS base class with time as
    DATA in `node_time` (the slicer derives `_Y` classes); m->e nodes are YEAR-FREE."""
    product_key = None
    for r in rows:
        if r.get("parameterCode") in LEVEL_OF_CODE:
            product_key = str(r[PRODUCT_KEY_COL])
            break

    # Structure comes from m-c and e-m. c-p is used only as a cross-check on
    # component mass, not as a structural edge: its inline material restates the
    # m-c row. Material/element node ids are scoped by the component path.
    mc_by, em_by, cp_total = _collect_facts(rows)

    # ---- masses (denominators), bottom-up ----
    mat_mass = {}                 # (comp, mat) -> material mass in its component
    for key, rs in mc_by.items():
        mat_mass[key], _ = aggregate_rows(rs)
    comp_mass = defaultdict(float)   # comp -> sum of its materials
    for (comp, mat), m in mat_mass.items():
        comp_mass[comp] += m
    car_mass = sum(comp_mass.values())

    # ---- cross-check c-p vs derived component mass (warn, don't skew) ----
    for comp, cpv in cp_total.items():
        cm = comp_mass.get(comp)
        if cm and abs(cpv - cm) > 1e-6 * max(1.0, cm):
            print(f"  [warn year {year}] component {comp[-1]}: c-p total "
                  f"{cpv:.4f} != sum(m-c) {cm:.4f}", file=sys.stderr)

    # ---- node names: product/component scoped by PRODUCT+YEAR; material/element by
    # PRODUCT ONLY (chemistry is year-invariant, so every year points at the SAME
    # material node — the year axis never reaches the m->e layer).
    tag = f"{product_key}_{year}"

    def comp_name(comp):
        return f"{comp[-1]}__{tag}"

    def mat_name(comp, mat):
        return f"{comp[-1]}_{mat[-1]}__{product_key}"

    def elem_name(comp, mat, sym):
        return f"{comp[-1]}_{mat[-1]}_{sym}__{product_key}"

    if scope is None:
        scope = {"year": int(year)}      # ad-hoc single-instance use
    # node name -> time scope ({year} | {start,end}): the ETL records production time
    # as DATA on the instance node and mints NO `_Y` slice class; the slicer derives
    # every time-slice class + edge from this referenceYear/period downstream.
    if node_time is None:
        node_time = {}

    car = f"{product_key}_{year}"
    # the instance is typed by its BASE class; its year lives in node_time.
    nodes = {car: ("Product", product_key)}
    node_time[car] = dict(scope)
    # itemMass (absolute kg) per Product/Component instance node — the reference
    # anchor turning kg/kg composition into absolute amounts (car_mass / comp_mass).
    # Material/Element nodes get none (matter is bulk, never an item).
    item_mass = {car: car_mass}
    statements = []

    def emit(whole, part, value, whole_mass, quality, uncertainty_limit=None):
        if not whole_mass:
            return
        best = value / whole_mass
        if best <= 0.0:
            return               # loader rejects zero-content statements
        # RECTANGULAR statement: best + DQ quality scores; half-width is DERIVED from
        # the DQ scores by the uncertaintyLimitStrategy. When a known-limit correction
        # applies, the override limit is stamped directly so the uncertainty plugin
        # uses it verbatim instead of re-deriving from the DQ band table.
        stmt = {
            "whole": whole, "part": part,
            "best": round(best, 9),
            "unit": "kgkg", "dist": "rectangular",
            "quality": _fill_quality(quality),
        }
        if uncertainty_limit is not None:
            stmt["uncertainty_limit"] = uncertainty_limit
        statements.append(stmt)

    def class_for(key, level, unknown_cls):
        """Resolve a component/material KEY to its futuram class. A catch-all key
        maps to the generic unknown filler; a real key must be a declared class else
        we fail loudly (a new key the ontology is missing)."""
        key = _KEY_ALIASES.get(key, key)   # known concept under another localname
        # a DECLARED class that is semantically a remainder bucket (elvRest, …)
        # maps to the unknown filler — checked BEFORE "declared class wins" so the
        # catch-all is never kept verbatim (mirrors otherOrUndefinedElements path).
        if key in _CATCHALL_CLASSES:
            return unknown_cls
        if key in ontology_classes:
            return key
        if is_catchall(key):
            return unknown_cls
        raise ValueError(f"{level} key {key!r} is not a futuram class and is not "
                         f"a recognised catch-all (year {year})")

    # ---- layer 1: car -> component (component mass / car mass) ----
    for comp, cm in comp_mass.items():
        comp_cls = class_for(comp[-1], "component", "unknownComponent")
        if _is_omit_class(comp_cls):
            continue                      # remainder — omit; resolver re-infers
        cn = comp_name(comp)
        # component instance typed by its timeless base class; its production
        # time is recorded as data so the plugin layer derives the slice.
        nodes[cn] = ("Component", comp_cls)
        node_time[cn] = dict(scope)
        item_mass[cn] = cm
        # car->component is a structural mass share; like every statement it carries
        # the full DQ vector (the default), and its rectangular spread is derived from
        # that by the strategy.
        emit(car, cn, cm, car_mass, None)

    # ---- layer 2: component -> material (material mass / component mass) ----
    for (comp, mat), m in mat_mass.items():
        mat_cls = class_for(mat[-1], "material", "unknownMaterial")
        if _is_omit_class(mat_cls):
            continue                      # remainder — omit; resolver re-infers
        cn = comp_name(comp)
        mn = mat_name(comp, mat)
        if cn not in nodes:
            continue                      # material of a component we omitted
        nodes[mn] = ("Material", mat_cls)
        _, q = aggregate_rows(mc_by[(comp, mat)])
        adj = _lc.lookup(corrections, comp[-1], mat[-1],
                         dq_validity=q.get("Validity"),
                         dq_accuracy=q.get("Accuracy"),
                         dq_consistency=q.get("Consistency"),
                         dq_integrity=q.get("Integrity"),
                         dq_timeliness=q.get("Timeliness"),
                         dq_completeness=q.get("Completeness"))
        if adj:
            q = {**q, **adj}
        emit(cn, mn, m, comp_mass[comp], q)

    # ---- layer 3: material -> element (year-invariant chemistry) ----
    # Emitted ONCE per year-free element node from the canonical (cross-year mean)
    # table; the statement describes every instance of the product (shared chemistry).
    if em_emitted is None:
        em_emitted = set()
    for (comp, mat, sym), rs in em_by.items():
        mn = mat_name(comp, mat)
        if mn not in nodes:
            continue                      # element of a material we didn't keep
        cls, _ = element_class(sym, ontology_classes)
        if cls is None or _is_omit_class(cls) or _is_omit_class(sym):
            continue                      # remainder element — omit; resolver re-infers
        en = elem_name(comp, mat, sym)
        nodes[en] = ("Element", cls)
        if en in em_emitted:
            continue                      # already emitted by an earlier year
        em_emitted.add(en)
        if em_canon is not None:
            entry = em_canon.get((product_key, comp, mat), {}).get(sym)
            if entry is None:
                continue                  # zero-content element: omitted
            best, q = entry
        else:
            val, q = aggregate_rows(rs)
            best = val
        adj = _lc.lookup(corrections, comp[-1], mat[-1],
                         dq_validity=q.get("Validity"),
                         dq_accuracy=q.get("Accuracy"),
                         dq_consistency=q.get("Consistency"),
                         dq_integrity=q.get("Integrity"),
                         dq_timeliness=q.get("Timeliness"),
                         dq_completeness=q.get("Completeness"))
        if adj:
            q = {**q, **adj}
        if em_canon is not None:
            emit(mn, en, best, 1.0, q)
        else:
            emit(mn, en, best, mat_mass[(comp, mat)], q)

    return nodes, statements, {"product_key": product_key,
                               "car_mass": car_mass,
                               "item_mass": item_mass,
                               "node_time": node_time}


# ---------------------------------------------------------------------------
# Top level: assemble all instances into one scenario YAML
# ---------------------------------------------------------------------------

def _row_year(r):
    y = r.get("productionYear")
    return int(float(y)) if y not in (None, "") else 0


# Relative tolerance for the validity-period dedup: two years are "the same
# composition" when every part of every statement agrees this closely (targets
# literally-duplicated rows; a genuine change is far above this).
DEDUP_TOL = 1e-9


def _row_quality(rs):
    """The DQV quality dict of a (whole,part) row group (per aggregate_rows)."""
    q = {}
    for col, dim in DQ_COLS.items():
        if rs[0].get(col) not in (None, ""):
            q[dim] = float(rs[0][col])
    return q


def _composition_fingerprint(rows):
    """A per-(whole,part) fingerprint of ONE (product, year)'s composition (the basis
    for validity-period dedup): best value + full DQV quality scores + (unit, dist).
    Returns {structural key -> {"v": (best,), "q": sorted items, "meta": (unit,dist)}}."""
    META = ("kgkg", "rectangular")   # mirrors emit()'s constants
    mc_by, em_by, cp_total = _collect_facts(rows)
    fp = {}
    for key, rs in mc_by.items():
        b, _ = aggregate_rows(rs)
        fp[("mc",) + tuple(key)] = {"v": (b,),
                                    "q": _q_key(_row_quality(rs)),
                                    "meta": META}
    for key, rs in em_by.items():
        b, _ = aggregate_rows(rs)
        fp[("em",) + tuple(key[0]) + tuple(key[1]) + (key[2],)] = {
            "v": (b,), "q": _q_key(_row_quality(rs)), "meta": META}
    for comp, total in cp_total.items():
        fp[("cp",) + tuple(comp)] = {"v": (total,), "q": (), "meta": META}
    return fp


def _q_key(q):
    return tuple(sorted(q.items()))


def _fp_equal(a, b, tol=DEDUP_TOL):
    """Two fingerprints are equal iff they cover the SAME (whole,part) keys and
    every key agrees on values (within tol), DQV quality and metadata. A
    statement present in one year but not the other blocks the merge."""
    if set(a) != set(b):
        return False
    for k, ea in a.items():
        eb = b[k]
        for x, y in zip(ea["v"], eb["v"]):
            if abs(x - y) > tol * max(1.0, abs(x), abs(y)):
                return False
        if ea["q"] != eb["q"] or ea["meta"] != eb["meta"]:
            return False
    return True


def _dedup_year_runs(by_inst):
    """Collapse each product's CONSECUTIVE years with identical composition
    into one run: returns {(prod, (y0, y1)): rows-of-y0}, plus a log line per
    collapsed run (no silent dedup). A run of one year keeps y0 == y1."""
    runs = {}
    by_prod = defaultdict(list)
    for (prod, yr) in sorted(by_inst):
        by_prod[prod].append(yr)
    for prod, years in by_prod.items():
        fps = {y: _composition_fingerprint(by_inst[(prod, y)]) for y in years}
        i = 0
        while i < len(years):
            j = i
            while (j + 1 < len(years) and years[j + 1] == years[j] + 1
                   and _fp_equal(fps[years[i]], fps[years[j + 1]])):
                j += 1
            runs[(prod, (years[i], years[j]))] = by_inst[(prod, years[i])]
            if j > i:
                print(f"  [dedup] {prod}: years {years[i]}..{years[j]} "
                      f"identical — collapsed into one period instance",
                      file=sys.stderr)
            i = j + 1
    return runs


def transform(path, sid=None, sheet=None, years=None, products=None):
    """Transform a dataset into a scenario doc, one INSTANCE per (product,
    productionYear). `years`/`products` optionally filter the slice; products not in
    the local hierarchy get a `subclass_of` edge to their productKeyLevel1 parent."""
    rows = read_rows(path, sheet=sheet)
    ontology_classes = ontology_class_set()
    corrections = _lc.load(path)

    # one instance per (product, year), after optional filtering
    by_inst = defaultdict(list)
    for r in rows:
        if r.get("parameterCode") not in LEVEL_OF_CODE:
            continue
        prod = str(r.get(PRODUCT_KEY_COL) or "")
        yr = _row_year(r)
        if years is not None and yr not in years:
            continue
        if products is not None and prod not in products:
            continue
        by_inst[(prod, yr)].append(r)
    if not by_inst:
        raise ValueError("no rows match the product/year filter")

    # validity-period dedup: consecutive identical years collapse into ONE
    # period instance (and ONE period-scoped slice class) instead of N
    # duplicates. Logged per run — never silent.
    runs = _dedup_year_runs(by_inst)
    insts = sorted(runs)

    # the year-invariant material->element layer: canonical cross-year table,
    # ENFORCED (raises beyond EM_YEAR_TOL). Each m->e statement is then emitted
    # once, on year-free nodes — undoing the per-year duplication of chemistry.
    em_canon = _em_canonical(by_inst)
    em_emitted = set()

    all_nodes = {}
    all_item_mass = {}                 # node name -> absolute item mass (kg)
    all_statements = []
    infos = []
    node_time = {}                     # instance node name -> time scope (DATA)
    parents = {}                       # productKeyLevel2 -> productKeyLevel1
    for prod, (y0, y1) in insts:
        rs = runs[(prod, (y0, y1))]
        scope = ({"year": y0} if y0 == y1 else {"start": y0, "end": y1})
        label = str(y0) if y0 == y1 else f"{y0}_{y1}"
        nodes, stmts, info = build_instance(rs, label, ontology_classes,
                                            em_canon=em_canon,
                                            em_emitted=em_emitted,
                                            scope=scope, node_time=node_time,
                                            corrections=corrections)
        all_nodes.update(nodes)
        all_item_mass.update(info.get("item_mass", {}))
        all_statements.extend(stmts)
        infos.append((prod, label, info))
        # remember the product's parent (level-1 key) for subclass_of
        p1 = rs[0].get("productKeyLevel1")
        if p1:
            parents[prod] = str(p1)

    # subclass edges: only the BASE-level taxonomy edge for products not declared in
    # the local hierarchy (under their level-1 parent). NO `_Y` slice edges — the
    # plugin layer derives those from the instances' referenceYear/period.
    products = {prod for prod, _ in insts}
    subclass_of = {}
    for p in products:
        if p in parents and p not in ontology_classes:
            subclass_of[p] = [parents[p]]

    # Rich rdfs:labels for the vehicle product CLASSES: decode the V-code into
    # drivetrain + body segment, so a segment-phrased question resolves to the right
    # class by label (the CSV has no description; identity lives in the digits).
    class_labels = {}
    for p in products:
        lbl = _product_label(p, parents.get(p))
        if lbl:
            class_labels[p] = lbl

    # the drivetrain axis VALUE marker (generic, source-derived): productKeyLevel1 is
    # the product's drivetrain (elvBEV/…). Mark each value-class so the generic
    # ValueAxisSlicer finds it FROM THE GRAPH; unknown* fillers are never values.
    axis_values = {dt: "drivetrain-mean" for p, dt in parents.items()
                   if dt and not str(dt).startswith("unknown")}

    r0 = rows[0]
    yrs = sorted({y for _, (y0, y1) in insts for y in (y0, y1)})
    provenance = {
        "source": str(r0.get("reference") or "FutuRaM dataset")[:120],
        "agent": str(r0.get("dataProcessorInstitution") or "FutuRaM"),
        "production": "use",
        "validFrom": f"{yrs[0]}-01-01",
        "validUntil": f"{yrs[-1]}-12-31",
    }

    sid = sid or path.stem.replace("-", "_").replace(".", "_")
    doc = {
        "id": sid,
        "title": f"transformed from {path.name}: {len(insts)} (product,year) "
                 f"instance(s), per-layer kg/kg",
        "note": _build_note(infos),
    }
    if subclass_of:
        doc["subclass_of"] = subclass_of
    if class_labels:
        doc["class_labels"] = class_labels   # V-class IRI -> rich rdfs:label
    if axis_values:
        doc["axis_values"] = axis_values     # class -> axis strategy token
    if node_time:
        doc["node_time"] = node_time         # instance node -> time scope (DATA)
    doc["provenance"] = provenance
    doc["nodes"] = {n: ({"level": lv, "class": cls,
                         "itemMass": round(all_item_mass[n], 6)}
                        if n in all_item_mass
                        else {"level": lv, "class": cls})
                    for n, (lv, cls) in all_nodes.items()}
    doc["statements"] = all_statements
    # NO doc-side slice derivation: time-slice classes (leaf AND ancestor) are derived
    # in the plugin/RDF layer (builder.slicer.YearSlicer) from referenceYear/period.
    # The ETL emits only base-typed instances + their time as data.
    return doc


def transform_doc(path, sid=None, sheet=None, years=None, products=None,
                  canonicalize=True):
    """transform() the dataset to a doc, optionally honest-material-canonicalised.
    The doc is the in-memory composition spec the RDF emitter consumes (no YAML
    written); carries nodes/statements/subclass_of/provenance + axis_values."""
    from . import chain_loader
    doc = transform(path, sid=sid, sheet=sheet, years=years, products=products)
    if canonicalize:
        doc, _changes = chain_loader.canonicalize_doc(doc)
    return doc


# the m->e layer's node-LEVEL signature: a statement is m->e iff its whole node is
# a Material and its part an Element (the chemistry layer is year-invariant and
# factored into its own sibling file).
_EM_WHOLE_LEVEL, _EM_PART_LEVEL = "Material", "Element"


def split_em_doc(doc):
    """Partition a transform doc into (structural_doc, em_doc) by node level: em_doc
    holds ONLY the year-invariant m->e statements + their nodes; structural_doc holds
    every OTHER statement and ALL nodes (element nodes stay for routing)."""
    level = {n: spec["level"] for n, spec in doc["nodes"].items()}

    def is_em(s):
        return (level.get(s["whole"]) == _EM_WHOLE_LEVEL
                and level.get(s["part"]) == _EM_PART_LEVEL)

    em_stmts = [s for s in doc["statements"] if is_em(s)]
    rest_stmts = [s for s in doc["statements"] if not is_em(s)]

    def _shell(extra):
        # node_time carries each instance's production time as DATA; it MUST propagate
        # into the split sub-docs or the bucket/em graphs lose all instance time and
        # the slicer mints no slices (no class_time).
        d = {k: doc[k] for k in
             ("title", "note", "subclass_of", "class_time", "node_time",
              "provenance", "axis_values") if k in doc}
        d.update(extra)
        return d

    structural = _shell({"id": doc.get("id"), "nodes": doc["nodes"],
                         "statements": rest_stmts})
    keep = {s["whole"] for s in em_stmts} | {s["part"] for s in em_stmts}
    em_nodes = {n: doc["nodes"][n] for n in sorted(keep) if n in doc["nodes"]}
    em = _shell({"id": (doc.get("id") + "_em") if doc.get("id") else "em",
                 "nodes": em_nodes, "statements": em_stmts})
    em.pop("axis_values", None)   # the chemistry layer carries no drivetrain notion
    return structural, em


def to_graph(path, sid=None, sheet=None, years=None, products=None,
             canonicalize=True, full_metadata=True):
    """Transform a dataset straight into composition-statement RDF (dataset -> doc ->
    composition_rdf, pure rdflib, NO YAML/oracle). Iso-equivalent to the frozen oracle
    serializer PLUS the generic drivetrain axis-value marker from productKeyLevel1."""
    from . import chain_loader
    from .composition_rdf import composition_rdf
    doc = transform(path, sid=sid, sheet=sheet, years=years, products=products)
    if canonicalize:
        doc, _changes = chain_loader.canonicalize_doc(doc)
    return composition_rdf(doc, full_metadata=full_metadata,
                           axis_values=doc.get("axis_values"))


def _build_note(infos):
    n_prod = len({p for p, _, _ in infos})
    lines = [
        "Auto-generated from a FutuRaM composition dataset (kg/unit absolute "
        "mass, converted to per-layer kg/kg). Each (product, productionYear) is "
        "a separate instance.",
        "",
        "Levels from parameterCode: c-p (component-of-product), m-c "
        "(material-of-component), e-m (element-of-material). Each value is "
        "divided by its IMMEDIATE whole's mass to give the per-layer fraction; "
        "component structure is taken from m-c (c-p is a cross-check, since its "
        "inline material duplicates the m-c row).",
        "",
        "The material->element layer is YEAR-INVARIANT (intrinsic chemistry, "
        "enforced within EM_YEAR_TOL): material/element nodes carry no year "
        "and each m->e statement appears once, shared by every production-year "
        "instance of its product.",
        "",
        f"{len(infos)} instance(s) across {n_prod} product class(es):",
    ]
    for prod, yr, info in infos:
        lines.append("  {} / {}: total mass {:.1f} kg".format(
            prod, yr, info["car_mass"] or 0.0))
    return "\n".join(lines)


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("input", type=pathlib.Path, help="Excel (.xlsx) or CSV file")
    ap.add_argument("-o", "--output", type=pathlib.Path, required=True,
                    help="output file. .ttl -> composition RDF (no YAML); "
                         "otherwise a scenario YAML")
    ap.add_argument("--id", dest="sid", default=None,
                    help="scenario id (default: input stem)")
    ap.add_argument("--sheet", default=None, help="Excel sheet name")
    ap.add_argument("--years", default=None,
                    help="year filter: 'YYYY' or 'YYYY-YYYY' (inclusive range)")
    ap.add_argument("--product", dest="products", action="append", default=None,
                    help="productKeyLevel2 to keep (repeatable; default all)")
    ap.add_argument("--no-canonicalize", action="store_true",
                    help="skip honest-material canonicalisation of the output")
    ap.add_argument("--served", dest="served", type=pathlib.Path, default=None,
                    help="also write the query-optimized fq: dataset to this .ttl")
    args = ap.parse_args(argv)

    years = None
    if args.years:
        if "-" in args.years:
            a, b = args.years.split("-")
            years = set(range(int(a), int(b) + 1))
        else:
            years = {int(args.years)}
    products = set(args.products) if args.products else None
    canon = not args.no_canonicalize

    is_rdf = str(args.output).lower().endswith((".ttl", ".turtle"))

    comp_graph = None
    if is_rdf:
        # RDF-direct: dataset -> doc -> composition RDF (composition_rdf), NO oracle,
        # NO YAML on disk. This is the real-data emitter.
        comp_graph = to_graph(args.input, sid=args.sid, sheet=args.sheet,
                              years=years, products=products, canonicalize=canon)
        comp_graph.serialize(destination=str(args.output), format="turtle")
        from rdflib import RDF
        from common.vocab import FUT
        n_stmts = sum(1 for _ in comp_graph.subjects(
            RDF.type, FUT.CompositionStatement))
        print(f"wrote {args.output}  ({len(comp_graph)} triples, "
              f"{n_stmts} statements)" + ("  [canonicalised]" if canon else ""))
    else:
        doc = transform(args.input, sid=args.sid, sheet=args.sheet,
                        years=years, products=products)
        with open(args.output, "w") as fh:
            yaml.safe_dump(doc, fh, sort_keys=False, default_flow_style=False,
                           width=1000, allow_unicode=True)
        print(f"wrote {args.output}  ({len(doc['nodes'])} nodes, "
              f"{len(doc['statements'])} statements, id={doc['id']})")
        if canon:
            from . import canonicalize_materials
            changes = canonicalize_materials.canonicalize(args.output)
            print(f"  canonicalised {len(changes)} material(s) to honest classes")

    # optionally also persist the served fq: query graph (the virtual ontology).
    # The builder is RDF-in/RDF-out: resolve the composition GRAPH directly (no
    # oracle SupplyChain) — for the YAML branch, emit the composition RDF first.
    if args.served is not None:
        from rdflib import Graph
        from builder import resolver
        from common import pipeline
        if comp_graph is None:
            comp_graph = to_graph(args.input, sid=args.sid, sheet=args.sheet,
                                  years=years, products=products,
                                  canonicalize=canon)
        sg = Graph()
        resolver.resolve_all(comp_graph, into=sg)
        sg.parse(str(pipeline.QUERY_TBOX), format="turtle")   # fq: TBox terms
        sg.serialize(destination=str(args.served), format="turtle")
        print(f"wrote {args.served}  (served fq: graph, {len(sg)} triples)")


if __name__ == "__main__":
    main()
